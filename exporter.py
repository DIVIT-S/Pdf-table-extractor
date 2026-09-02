import io
import logging
from typing import List, Dict, Any
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def consolidate_tables(tables: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Merge tables sharing identical column counts and headers across pages."""
    groups: List[Dict[str, Any]] = []
    for t in tables:
        df: pd.DataFrame = t.get("df")
        if df is None or df.empty:
            continue
        hdr, cols = tuple(str(c).strip().lower() for c in df.iloc[0]), df.shape[1]
        for g in groups:
            if g["cols"] == cols and g["hdr"] == hdr:
                if df.shape[0] > 1:
                    d = df.iloc[1:].copy()
                    d.columns = g["df"].columns
                    g["df"] = pd.concat([g["df"], d], ignore_index=True)
                if t.get("page") and t.get("page") not in g["pages"]:
                    g["pages"].append(t.get("page"))
                break
        else:
            groups.append({
                "table_num": len(groups) + 1, "hdr": hdr, "cols": cols,
                "df": df.copy().reset_index(drop=True),
                "pages": [t.get("page")] if t.get("page") else [1],
            })
    return groups


def export_tables_to_excel(tables: List[pd.DataFrame]) -> bytes:
    """
    Compile a list of pandas DataFrames into a formatted in-memory .xlsx workbook.
    All DataFrames are saved sequentially into a single sheet ("All_Tables").
    Uses openpyxl directly to avoid pandas/openpyxl integration issues.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    import re
    
    def sanitize_cell_value(val):
        """Remove invalid XML characters and ensure safe cell value."""
        if val is None or pd.isna(val):
            return ""
        
        val_str = str(val).strip()
        
        # Remove control characters (ASCII 0-31 except tab, newline, carriage return)
        val_str = ''.join(c for c in val_str if ord(c) >= 32 or c in '\t\n\r')
        
        # Remove other problematic characters
        val_str = val_str.replace('\x00', '')  # Null bytes
        val_str = val_str.replace('\ufffe', '')  # Byte Order Mark (BOM)
        val_str = val_str.replace('\ufeff', '')  # Zero-width No-Break Space
        
        # Replace actual newlines with spaces for Excel compatibility
        val_str = val_str.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')
        
        return val_str
    
    wb = Workbook()
    ws = wb.active
    ws.title = "All_Tables"
    
    if not tables:
        ws["A1"] = "No bordered tables were detected in this document."
    else:
        current_row = 1
        
        for idx, df in enumerate(tables, start=1):
            # Write table title
            title_cell = ws.cell(row=current_row, column=1, value=f"Table {idx}")
            title_cell.font = Font(name="Calibri", size=13, bold=True)
            title_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Merge cells for title if table has multiple columns
            if df.shape[1] > 1:
                ws.merge_cells(start_row=current_row, start_column=1, 
                               end_row=current_row, end_column=df.shape[1])
            
            current_row += 1
            
            # Define styles
            header_font = Font(name="Calibri", size=11, bold=True, color="000000")
            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            thin_border = Border(
                left=Side(style="thin", color="D3D3D3"),
                right=Side(style="thin", color="D3D3D3"),
                top=Side(style="thin", color="D3D3D3"),
                bottom=Side(style="thin", color="D3D3D3")
            )
            cell_font = Font(name="Calibri", size=11)
            align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Write header row (first row of DataFrame)
            for col_idx in range(df.shape[1]):
                header_val = sanitize_cell_value(df.iloc[0, col_idx])
                cell = ws.cell(row=current_row, column=col_idx + 1, value=header_val)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_left
                cell.border = thin_border
            
            current_row += 1
            
            # Write data rows (rows 1+ of DataFrame)
            for row_idx in range(1, df.shape[0]):
                for col_idx in range(df.shape[1]):
                    cell_val = sanitize_cell_value(df.iloc[row_idx, col_idx])
                    cell = ws.cell(row=current_row, column=col_idx + 1, value=cell_val)
                    cell.font = cell_font
                    cell.alignment = align_left
                    cell.border = thin_border
                
                current_row += 1
            
            # Add blank row after table
            current_row += 1
        
        # Auto-adjust column widths
        for col_idx in range(1, df.shape[1] + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            
            for row_idx in range(1, current_row):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    val_str = str(cell.value)
                    lines = val_str.split("\n")
                    cell_len = max(len(line) for line in lines) if lines else 0
                    max_len = max(max_len, cell_len)
            
            adjusted_width = min(max(max_len + 3, 10), 60)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save to bytes
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer.getvalue()