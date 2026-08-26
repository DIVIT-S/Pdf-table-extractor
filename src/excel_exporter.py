import io
import logging
from typing import List
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def export_tables_to_excel(tables: List[pd.DataFrame]) -> bytes:
    """
    Compile a list of pandas DataFrames into a formatted in-memory .xlsx workbook.
    Each DataFrame is saved to a distinct sheet (e.g. Table_1, Table_2).
    """
    output_buffer = io.BytesIO()

    with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
        if not tables:
            empty_df = pd.DataFrame({"Status": ["No tables detected in the uploaded PDF document."]})
            empty_df.to_excel(writer, sheet_name="Summary", index=False)
        else:
            for idx, df in enumerate(tables, start=1):
                sheet_name = f"Table_{idx}"[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Access the openpyxl worksheet to apply styling
                ws = writer.sheets[sheet_name]

                header_font = Font(name="Calibri", size=11, bold=True, color="000000")
                header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                thin_border = Border(
                    left=Side(style="thin", color="D3D3D3"),
                    right=Side(style="thin", color="D3D3D3"),
                    top=Side(style="thin", color="D3D3D3"),
                    bottom=Side(style="thin", color="D3D3D3"),
                )
                cell_font = Font(name="Calibri", size=11)
                align_left = Alignment(horizontal="left", vertical="center", wrap_text=False)

                # Style header row
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_left
                    cell.border = thin_border

                # Style data rows & calculate column widths
                for col_idx, col in enumerate(ws.columns, start=1):
                    max_len = 0
                    col_letter = get_column_letter(col_idx)

                    for cell in col:
                        cell.border = thin_border
                        if cell.row != 1:
                            cell.font = cell_font
                            cell.alignment = align_left

                        val_str = str(cell.value or "")
                        if len(val_str) > max_len:
                            max_len = len(val_str)

                    # Set adaptive column width (min 10, max 60)
                    adjusted_width = max(max_len + 3, 10)
                    adjusted_width = min(adjusted_width, 60)
                    ws.column_dimensions[col_letter].width = adjusted_width

    output_buffer.seek(0)
    return output_buffer.getvalue()
